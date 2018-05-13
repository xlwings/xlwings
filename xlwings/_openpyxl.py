try:
    import openpyxl
except ImportError:
    openpyxl = None

import os, os.path
import numbers
import datetime as dt
import logging
logger = logging.getLogger(__name__)


class Engine(object):

    def __init__(self):
        self.apps = Apps()

    @property
    def name(self):
        return "Openpyxl"


def _clean_value_data_element(value, datetime_builder, empty_as, number_builder):
    if value == '':
        return empty_as
    if isinstance(value, dt.datetime) and datetime_builder is not dt.datetime:
        value = datetime_builder(
            month=value.month,
            day=value.day,
            year=value.year,
            hour=value.hour,
            minute=value.minute,
            second=value.second,
            microsecond=value.microsecond,
            tzinfo=None
        )
    elif number_builder is not None and type(value) == float:
        value = number_builder(value)
    return value


def clean_value_data(data, datetime_builder, empty_as, number_builder):
    return [[_clean_value_data_element(c, datetime_builder, empty_as, number_builder) for c in row] for row in data]

Engine.clean_value_data = staticmethod(clean_value_data)


def prepare_xl_data_element(x):
    return x

Engine.prepare_xl_data_element = staticmethod(prepare_xl_data_element)


class Apps(object):

    def __init__(self):
        self._apps = [App(self)]

    def __iter__(self):
        return iter(self._apps)

    def __len__(self):
        return len(self._apps)

    def __getitem__(self, index):
        return self._apps[index]

    def add(self, **kwargs):
        self._apps.insert(0, App(self, **kwargs))
        return self._apps[0]

class App(object):

    _next_pid = -1

    def __init__(self, apps, add_book=True, **kwargs):
        self.apps = apps
        self._pid = App._next_pid
        App._next_pid -= 1
        self._books = Books(self)
        if add_book:
            self._books.add()

    @property
    def engine(self):
        return engine

    def activate(self, steal_focus=False):
        raise NotImplementedError()

    @property
    def books(self):
        return self._books

    @property
    def pid(self):
        return self._pid

    def kill(self):
        self.apps._apps.remove(self)
        self.apps = None

    def range(self, arg1, arg2=None):
        # TODO: better implementation
        return self.books.active.sheets.active.range(arg1)



class Books(object):

    def __init__(self, app):
        self.app = app
        self.books = []
        self.active = None

    def open(self, filename):
        filename = os.path.abspath(filename)
        book = self._try_find_book_by_name(filename)
        if book is None:
            book = Book(
                api=openpyxl.load_workbook(filename), 
                books=self,
                path=filename
            )
            self.books.append(book)
        self.active = book
        return book

    @property
    def api(self):
        return None

    def add(self):
        name = "Book" + str(len(self.books) + 1)
        book = Book(
            api=openpyxl.Workbook(),
            books=self,
            path=name
        )
        self.books.append(book)
        self.active = book
        return book

    def _try_find_book_by_name(self, name):
        for book in self.books:
            if (book.name == name
                or book.fullname == name):
                return book
        return None

    def __call__(self, name_or_index):
        if isinstance(name_or_index, numbers.Number):
            return self.books[name_or_index - 1]
        else:
            book = self._try_find_book_by_name(name_or_index)
            if book is None:
                raise KeyError(name_or_index)
            return book


    def __len__(self):
        return len(self.books)

    def __iter__(self):
        for book in self.books:
            yield book


class Book(object):

    def __init__(self, api, books, path):
        self.api = api
        self.path = path
        self.books = books

    @property
    def name(self):
        _, n = os.path.split(self.path)
        return n

    @property
    def fullname(self):
        # TODO: implement correctly
        return self.path

    @property
    def sheets(self):
        return Sheets(book=self)

    @property
    def app(self):
        return self.books.app

    @property
    def index(self):
        return self.app.books.index(self)

    def close(self):
        assert self.api is not None, "Seems this book was already closed."
        self.books.books.remove(self)
        self.books = None
        self.api = None
        self.filename = None

    def save(self, path=None):
        if self.path:
            del self.app.books[self.fullname]
            self.filename = path
            self.app.books[self.filename] = self
        self.api.save(self.filename)


class Sheets(object):

    def __init__(self, book):
        self.book = book

    @property
    def api(self):
        return None

    @property
    def active(self):
        return Sheet(
            api=self.book.api.active,
            sheets=self
        )

    def __call__(self, name_or_index):
        if isinstance(name_or_index, int):
            api = self.book.api.worksheets[name_or_index - 1]
        else:
            api = self.book.api[name_or_index]
        return Sheet(api=api, sheets=self)

    def __len__(self):
        return len(self.book.api.worksheets)

    def __iter__(self):
        for sheet in self.book.api.worksheets:
            yield Sheet(api=sheet, sheets=self)

    def add(self, before=None, after=None):
        return Sheet(
            api=self.book.api.create_sheet(),
            sheets=self
        )


class Sheet(object):

    def __init__(self, api, sheets):
        self.api = api
        self.sheets = sheets

    @property
    def name(self):
        return self.api.title

    @name.setter
    def name(self, value):
        self.api.title = value

    @property
    def book(self):
        return self.sheets.book

    @property
    def index(self):
        return self.sheets.book.index(self.api)

    def range(self, arg1, arg2=None):
        if isinstance(arg1, Range):
            arg1 = arg1.api[0][0].coordinate
        if isinstance(arg2, Range):
            arg2 = arg2.api[-1][-1].coordinate
        if arg2 is None:
            api = self.api[arg1]
        else:
            api = self.api["%s:%s" % (arg1, arg2)]
        return Range(api=api, sheet=self)

    def activate(self):
        self.sheets.book.api.active_sheet = self.sheets.book.api.index(self.api)

    def select(self):
        self.sheets.book.api.active_sheet = self.sheets.book.api.index(self.api)

    def clear(self):
        return NotImplementedError()

    def autofit(self, axis=None):
        logger.warning("Autofit doesn't do anything in openpyxl engine.")


class Range(object):

    def __init__(self, api, sheet):
        if isinstance(api, tuple):
            self.api = api
        else:
            self.api = ((api,),)
        self.sheet = sheet

    def coords(self):
        return (
            self.sheet.name,
            self.row,
            self.column,
            len(self.api),
            len(self.api[0])
        )

    def __len__(self):
        return len(self.api) * len(self.api[0])

    @property
    def row(self):
        return self.api[0][0].row_idx

    @property
    def column(self):
        return self.api[0][0].col_idx

    @property
    def shape(self):
        return len(self.api), len(self.api[0])

    @property
    def raw_value(self):
        if len(self.api) == 1 and len(self.api[0]) == 1:
            return self.api[0][0].value
        else:
            return tuple(
                tuple(
                    cell.value for cell in row
                )
                for row in self.api
            )

    @raw_value.setter
    def raw_value(self, value):
        if isinstance(value, tuple) or isinstance(value, list):
            if (len(value), len(value[0])) != self.shape:
                assert False, "Not implemented"
            for i in range(len(value)):
                for j in range(len(value[0])):
                    self.api[i][j].value = value[i][j]
        else:
            for row in self.api:
                for cell in row:
                    cell.value = value

    @property
    def address(self):
        return self.api[0][0].coordinate

    def __call__(self, row, col):
        # TODO: better implementation
        return Range(
            api=self.api[0][0].offset(row-1, col-1),
            sheet=self.sheet
        )

engine = Engine()
